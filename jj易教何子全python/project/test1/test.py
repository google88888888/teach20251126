import pandas as pd
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
# data = {
#     '员工ID': [1001, 1002, 1003, 1004, 1005],
#     '姓名': ['张三', '李四', '王五', '赵六', '钱七'],
#     '部门': ['技术部', '销售部', '技术部', '人事部', '销售部'],
#     '入职日期': ['2020-01-15', '2019-03-20', '2021-07-10', '2018-11-05', '2022-02-28'],
#     '基本工资': [8000, 6000, 8500, 5500, 6500],
#     '奖金': [2000, 3000, 2500, 1500, 2800],
#     '绩效评分': [85, 92, 78, 88, 95]
# }

# df = pd.DataFrame(data)
# df.to_excel('员工数据.xlsx', index=False, sheet_name='员工信息')

# dfRead = pd.read_excel('科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103.xlsx', sheet_name='科目余额表',header=[0,1,2,3])
# print(dfRead)
# print('aaaaaaaaaaaaaaaaaaaaaaaaaa')
# print(list(dfRead.columns))
# print('aaaaaaaaaaaaaaaaaaaaaaaaaa')
# print(dfRead.values)
# print('aaaaaaaaaaaaaaaaaaaaaaaaaa')
# print(dfRead.T.values)
# print('aaaaaaaaaaaaaaaaaaaaaaaaaa')
# originColums=list(dfRead.columns)
# newList=dfRead.T.values.tolist()


# with open('originColums.json', 'w', encoding='utf-8') as f:
#     json.dump(originColums, f, ensure_ascii=False, indent=4)





# with open('newList.json', 'w', encoding='utf-8') as f:
#     json.dump(newList, f, ensure_ascii=False, indent=4)


# dfRead.iat[0, 0] = dfRead.iat[0, 0]+'ffffffffffff'
# dfRead.to_excel('测试输出.xlsx', sheet_name='员工信息')

# # 使用openpyxl加载原始文件以保留格式
# wb_original = load_workbook('科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103.xlsx')
# ws_original = wb_original.active

# # 加载临时文件获取修改后的数据
# wb_temp = load_workbook('测试输出.xlsx')
# ws_temp = wb_temp.active

# 只更新数据区域的值，保留表头格式


# 保存最终文件
# wb_original.save('测试输出1.xlsx')




# with pd.ExcelWriter('./科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103.xls', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#     dfRead.to_excel(writer, sheet_name='科目余额表', index=False)

# newList[1][10]='fffffffffffffffffffff'
# newData={}
# for index,item in enumerate(originColums):
#     newData[item]=newList[index]



# dfOut = pd.DataFrame(newData)
# dfOut.to_excel('测试输出.xlsx', index=False, sheet_name='员工信息')



# a = [1006 ,'1' ,'2' ,'2023-04-05', 6 ,7 ,8]
# for index,item in enumerate(newList):
#     item.append(a[index])
#     print(index,item)
# newData={}
# for index,item in enumerate(originColums):
#     newData[item]=newList[index]
# print(newData)


# dfOut = pd.DataFrame(newData)
# dfOut.to_excel('员工数据1.xlsx', index=False, sheet_name='员工信息')


# 使用xlwings直接操作Excel，保留所有格式


def count_leading_spaces(s):
    count = 0
    for char in s:
        if char == ' ':
            count += 1
        else:
            break
    return count

dfRead = pd.read_excel('科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103.xls', sheet_name='科目余额表',header=[0,1,2,3],skipfooter=2)
newList=dfRead.T.values.tolist()[1]
prefix=[]
outputList=[]
for item in newList:
    prefixSpaces=count_leading_spaces(item)//2
    itemWithoutSpaces=item[prefixSpaces*2:]
    if prefixSpaces==0:
        prefix=[itemWithoutSpaces]
    else:
        if len(prefix)>prefixSpaces:
            prefix[prefixSpaces]=prefix[prefixSpaces-1]+"-"+itemWithoutSpaces
        else:
            prefix.append(prefix[prefixSpaces-1]+"-"+itemWithoutSpaces)
    if len(prefix)==1:
        outputList.append(itemWithoutSpaces)
    else:
        outputList.append(prefix[prefixSpaces-1]+"-"+itemWithoutSpaces)
print(outputList)
        
with open('newList1.json', 'w', encoding='utf-8') as f:
    json.dump(outputList, f, ensure_ascii=False, indent=2)

app = xw.App(visible=False)  # 不显示Excel界面
wb = app.books.open('科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103.xls')
ws = wb.sheets[0]

for index,item in enumerate(outputList):
    ws[4+index, 1].value=item
# print('ws11111111111111111111111')
# print(ws[5, 1].value)
# print(ws[6, 1].value)
# print('ws2222222222222222222222222')
# ws[10, 1].value = ws[10, 1].value+'fffffffffffffff'


wb.save('科目余额表_2025年9月-2025年9月_广州白云恒利包装材料有限公司_20251103_modified7.xls')
wb.close()
app.quit()


